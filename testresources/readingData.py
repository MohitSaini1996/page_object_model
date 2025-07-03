import xlrd
import openpyxl
from openpyxl.reader.excel import load_workbook




class XLSReader:
    def __init__(self,path):
        self.path = path
        self.workbook = load_workbook(self.path)

    def getCellData(self,sheetname,rowIndex,colIndex):
        sheet = self.workbook[sheetname]
        return sheet.cell(rowIndex+1,colIndex+1).value

    def getCellDatabyColName(self,sheetname,rowIndex,colname):
        sheet = self.workbook[sheetname]
        for cnum in range(1,self.getColumnCount(sheetname)+1):
            extractedcolname = sheet.cell(1, cnum).value
            if extractedcolname==colname:
                celdata = sheet.cell(rowIndex+1,cnum).value
                return celdata
        return None


    def checkemptycell(self,sheetname,rowIndex,colIndex):
        sheet = self.workbook[sheetname]
        cell_value = sheet.cell(row=rowIndex+1,column=colIndex+1).value
        if cell_value is None:
            return True
        elif isinstance(cell_value,str) and cell_value.strip() == "":
            return True
        else:
            return False


    def getRowCount(self,sheetname):
        sheet = self.workbook[sheetname]
        return sheet.max_row

    def getColumnCount(self,sheetname):
        sheet = self.workbook[sheetname]
        return sheet.max_column